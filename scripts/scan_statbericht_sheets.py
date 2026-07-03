# -*- coding: utf-8 -*-
"""Scannt alle csv-* Blätter der Statistischen Berichte: #Bundesländer, Abschluss-/Schulart-Spalte, Jahre."""
import openpyxl, os, warnings
warnings.filterwarnings("ignore")
RAW = r"C:\Users\maxot\OneDrive\Desktop\datastory_school\data\raw"
files = ["statbericht_allgbild_2022-23.xlsx", "statbericht_allgbild_2023-24.xlsx"]

for fn in files:
    p = os.path.join(RAW, fn)
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    print("="*95)
    print(f"DATEI: {fn}")
    for s in [x for x in wb.sheetnames if x.startswith("csv-")]:
        ws = wb[s]
        rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        header = rows[0]
        def has(name): return any(name.lower() in c.lower() for c in header)
        iL = next((i for i,c in enumerate(header) if "bundesland" in c.lower()), None)
        nL = 0
        if iL is not None:
            nL = len(set(r[iL] for r in rows[1:] if iL < len(r) and r[iL] != ""))
        flags = []
        if has("Abschluss"): flags.append("Abschluss")
        if has("Schulart"): flags.append("Schulart")
        if has("Foerderschwerpunkt"): flags.append("Foerder")
        if has("Geschlecht"): flags.append("Geschl")
        # kurze Header-Vorschau
        hp = " | ".join([c for c in header if c][:7])
        print(f"  {s:16s} Zeilen={len(rows):5d} #BL={nL:2d}  [{','.join(flags)}]")
        print(f"       Header: {hp[:120]}")
