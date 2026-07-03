# -*- coding: utf-8 -*-
"""Exportiert die csv-* Blätter des Ausgaben-XLSX als eigenständige UTF-8 CSVs."""
import openpyxl, csv, os
P = r"C:\Users\maxot\OneDrive\Desktop\datastory_school\data\raw\21711_ausgaben_je_schueler_2024.xlsx"
OUT = r"C:\Users\maxot\OneDrive\Desktop\datastory_school\data\raw"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
sheets = [s for s in wb.sheetnames if s.startswith("csv-")]
for s in sheets:
    ws = wb[s]
    fn = os.path.join(OUT, f"ausgaben_{s.replace('csv-','')}.csv")
    n = 0
    with open(fn, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                w.writerow(["" if c is None else c for c in row])
                n += 1
    print(f"{s} -> {os.path.basename(fn)}  ({n} Zeilen)")
print("OK")
