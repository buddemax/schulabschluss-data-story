# -*- coding: utf-8 -*-
"""Inspiziert das Ausgaben-je-Schueler XLSX: Blätter, Dimensionen, erste Zeilen."""
import openpyxl
P = r"C:\Users\maxot\OneDrive\Desktop\datastory_school\data\raw\21711_ausgaben_je_schueler_2024.xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("BLAETTER:", wb.sheetnames)
for ws in wb.worksheets:
    print("="*80)
    print(f"BLATT: {ws.title}  (max_row~{ws.max_row}, max_col~{ws.max_column})")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [("" if c is None else str(c))[:22] for c in row]
        # nur nicht-leere Zeilen, erste 18
        if any(cells):
            print(f"  {i:2d}| " + " | ".join(cells[:9]))
        if i >= 17:
            break
